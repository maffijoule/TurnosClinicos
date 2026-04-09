"""
CEM — Módulo Hospitalizado v4
Horarios de turno fijos como input del usuario (2 sliders de cambio).
El solver decide: qué turno hace cada persona y qué días trabaja.
"""
import pulp, traceback
from datetime import date, timedelta
from flask import Blueprint, request, jsonify

hosp_bp = Blueprint('hosp', __name__)

DOW_MAP  = {0:'lunes',1:'martes',2:'miércoles',3:'jueves',4:'viernes',5:'sábado',6:'domingo'}
DIAS_LAB = ['lunes','martes','miércoles','jueves','viernes','sábado']
T_OPTS   = ['M','I','T']

def slot_str(s):
    return f"{(s//2)%24:02d}:{(s%2)*30:02d}"

def parse_slot(t):
    h, m = map(int, t.strip().split(':'))
    return h*2 + m//30

def build_semanas(mes, anio):
    primer = date(anio, mes, 1)
    ultimo = date(anio, mes+1 if mes<12 else 1, 1) - timedelta(days=1)
    semanas, sem, cur = [], [], primer
    while cur <= ultimo:
        if cur.month == mes:
            sem.append((cur.isoformat(), DOW_MAP[cur.weekday()]))
        cur += timedelta(days=1)
        if cur.weekday() == 0 and sem:
            semanas.append(sem); sem = []
    if sem: semanas.append(sem)
    return [s for s in semanas if any(d != 'domingo' for _,d in s)]

_cnt = [0]
def si(s):
    s = str(s)
    if not hasattr(si,'_cache'): si._cache = {}
    if s not in si._cache:
        si._cache[s] = str(_cnt[0]); _cnt[0] += 1
    return si._cache[s]

def resolver_hosp(data):
    si._cache = {}; _cnt[0] = 0
    import sys

    personal = data['personal']
    demanda  = data['demanda']
    cfg      = data.get('configuracion', {})

    mes       = int(cfg.get('mes',  3))
    anio      = int(cfg.get('anio', 2025))
    h_max_sem = float(cfg.get('horas_max_semana', 42))

    # ── Horarios de turno: dividir rango en 3 partes iguales ────────────────
    AP = parse_slot(cfg.get('hora_apertura', '07:00'))
    CI = parse_slot(cfg.get('hora_cierre',   '24:00'))
    rango = CI - AP
    # Dividir en 3 franjas iguales (redondear a slots enteros)
    dur1 = rango // 3
    dur2 = rango // 3
    dur3 = rango - dur1 - dur2   # absorbe el residuo

    TURNOS = {
        'M': {'ini': AP,          'fin': AP + dur1},
        'I': {'ini': AP + dur1,   'fin': AP + dur1 + dur2},
        'T': {'ini': AP + dur1 + dur2, 'fin': CI},
    }
    # Horas netas por turno por día (descuenta 30min colación)
    def h_turno(t):
        dur = TURNOS[t]['fin'] - TURNOS[t]['ini']
        return max((dur - 1) * 0.5, 0)   # -1 slot colación

    # Sábado: solo turno M opera
    DIAS_X_TURNO = {
        'M': DIAS_LAB,
        'I': ['lunes','martes','miércoles','jueves','viernes'],
        'T': ['lunes','martes','miércoles','jueves','viernes'],
    }

    print(f"[HOSP] Turnos: " +
          " | ".join(f"{t}:{slot_str(TURNOS[t]['ini'])}–{slot_str(TURNOS[t]['fin'])} ({h_turno(t)}h)"
                     for t in T_OPTS),
          flush=True, file=sys.stderr)

    semanas      = build_semanas(mes, anio)
    h_max        = {p['nombre']: float(p.get('horas_semana', h_max_sem)) for p in personal}
    eu_nombres   = [p['nombre'] for p in personal if p['rol'] == 'enfermera_eu']
    aseo_nombres = [p['nombre'] for p in personal if p['rol'] == 'aux_aseo']

    # ── Demanda → req_slot ────────────────────────────────────────────────────
    req_slot = {}
    for franja in demanda:
        ini = parse_slot(franja['inicio'])
        fin = parse_slot(franja['fin'])
        pab = int(franja['pabellones'])
        for s in range(ini, min(fin, 48)):
            req_slot[s] = max(req_slot.get(s, 0), pab)
    slots_activos = sorted(req_slot.keys())

    # Precomputar qué turno cubre cada slot activo
    def cubre(t, s):
        return TURNOS[t]['ini'] <= s < TURNOS[t]['fin']

    print(f"[HOSP] {mes}/{anio} | {len(personal)} pers | {len(slots_activos)} slots",
          flush=True, file=sys.stderr)

    # ── Modelo ────────────────────────────────────────────────────────────────
    prob = pulp.LpProblem("HOSP_v4", pulp.LpMinimize)

    # X[n,t,dow] = 1 si persona n trabaja turno t el día dow
    X = {}
    for p in personal:
        n = p['nombre']
        for t in T_OPTS:
            for dow in DIAS_X_TURNO[t]:
                X[n,t,dow] = pulp.LpVariable(f"X_{si(n)}_{t}_{si(dow)}", cat='Binary')

    # TW[n,t] = 1 si persona n tiene asignado turno t esta semana
    TW = {(n,t): pulp.LpVariable(f"TW_{si(n)}_{t}", cat='Binary')
          for p in personal for n in [p['nombre']] for t in T_OPTS}

    # Slack de cobertura
    SLK_eu   = {s: pulp.LpVariable(f"SEU_{s}",  lowBound=0) for s in slots_activos}
    SLK_aseo = {s: pulp.LpVariable(f"SAS_{s}",  lowBound=0) for s in slots_activos}

    # ── Objetivo: cobertura primero, personal segundo ─────────────────────────
    # Peso personal = 0.001 → máx aporte = 0.001 * N_dias << 1 slot de déficit
    PESO = 0.001
    prob += (pulp.lpSum(SLK_eu.values()) +
             pulp.lpSum(SLK_aseo.values()) +
             PESO * pulp.lpSum(X.values())), "Obj"

    # ── C1: max 1 turno/día/persona ───────────────────────────────────────────
    for p in personal:
        n = p['nombre']
        for dow in DIAS_LAB:
            vars_dia = [X[n,t,dow] for t in T_OPTS if (n,t,dow) in X]
            if vars_dia:
                prob += pulp.lpSum(vars_dia) <= 1, f"MaxTurno_{si(n)}_{si(dow)}"

    # ── C2: enlace X→TW (para C4) ────────────────────────────────────────────
    # TW[n,t]=1 si persona trabaja turno t algún día (no fija el turno toda la semana)
    for p in personal:
        n = p['nombre']
        for t in T_OPTS:
            for dow in DIAS_X_TURNO[t]:
                if (n,t,dow) in X:
                    prob += X[n,t,dow] <= TW[n,t], f"Cat_{si(n)}_{t}_{si(dow)}"

    # ── C2b: descanso mínimo 8h entre días consecutivos ───────────────────────
    # Si fin(t1) + 16 slots > ini(t2) + 48 (cruce medianoche) → no pueden combinarse
    CONSEC = [('lunes','martes'), ('martes','miércoles'), ('miércoles','jueves'),
              ('jueves','viernes'), ('viernes','sábado')]
    for p in personal:
        n = p['nombre']
        for d1, d2 in CONSEC:
            for t1 in T_OPTS:
                for t2 in T_OPTS:
                    if (n,t1,d1) not in X or (n,t2,d2) not in X:
                        continue
                    # slots de descanso = tiempo desde fin(t1) hasta ini(t2) cruzando medianoche
                    rest = TURNOS[t2]['ini'] + 48 - TURNOS[t1]['fin']
                    if rest < 16:   # menos de 8 horas
                        prob += X[n,t1,d1] + X[n,t2,d2] <= 1, \
                            f"Rest8h_{si(n)}_{t1}_{si(d1)}_{t2}_{si(d2)}"

    # ── C3: máx 42h semanales ────────────────────────────────────────────────
    # Horas = sum_días X[n,t,d] * h_turno(t)  — lineal porque h_turno es constante
    for p in personal:
        n = p['nombre']
        expr = [h_turno(t) * X[n,t,dow]
                for t in T_OPTS for dow in DIAS_X_TURNO[t]
                if (n,t,dow) in X]
        if expr:
            prob += pulp.lpSum(expr) <= h_max[n], f"HMax_{si(n)}"

    # ── C4: al menos 1 persona de cada rol en cada turno ─────────────────────
    # Garantiza cobertura en toda la jornada
    for rol, nombres_rol in [('enfermera_eu', eu_nombres), ('aux_aseo', aseo_nombres)]:
        if len(nombres_rol) >= len(T_OPTS):
            for t in T_OPTS:
                prob += pulp.lpSum(TW[n,t] for n in nombres_rol) >= 1, \
                    f"MinTurno_{rol}_{t}"

    # ── C5: cobertura EU por slot ─────────────────────────────────────────────
    for s in slots_activos:
        req_eu = max(1, -(-req_slot[s]//2))
        for dow in DIAS_LAB:
            cob = [X[n,t,dow] for n in eu_nombres for t in T_OPTS
                   if (n,t,dow) in X and cubre(t,s)]
            if cob:
                prob += pulp.lpSum(cob) + SLK_eu[s] >= req_eu, \
                    f"CovEU_{s}_{si(dow)}"

    # ── C6: cobertura aseo por slot ───────────────────────────────────────────
    for s in slots_activos:
        for dow in DIAS_LAB:
            cob = [X[n,t,dow] for n in aseo_nombres for t in T_OPTS
                   if (n,t,dow) in X and cubre(t,s)]
            if cob:
                prob += pulp.lpSum(cob) + SLK_aseo[s] >= 1, \
                    f"CovAseo_{s}_{si(dow)}"

    # ── Solve ─────────────────────────────────────────────────────────────────
    print(f"[HOSP] {len(prob.variables())} vars, {len(prob.constraints)} cons",
          flush=True, file=sys.stderr)
    prob.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=60))
    v      = pulp.value
    status = pulp.LpStatus[prob.status]
    print(f"[HOSP] Status: {status}", flush=True, file=sys.stderr)

    if not any(v(X[k]) is not None for k in X) or \
       prob.status == pulp.constants.LpStatusInfeasible:
        return {'status':'infeasible', 'mensaje':f'Sin solución ({status})'}

    # ── Extraer resultados ────────────────────────────────────────────────────
    semana_tipo = {}
    for p in personal:
        n = p['nombre']
        semana_tipo[n] = {'domingo': 'DOM'}
        for dow in DIAS_LAB:
            asig = next((t for t in T_OPTS
                         if (n,t,dow) in X and (v(X[n,t,dow]) or 0) > 0.5), '')
            semana_tipo[n][dow] = asig

    horas_tipo = {}
    for p in personal:
        n = p['nombre']
        h = sum(h_turno(t) for t in T_OPTS for dow in DIAS_X_TURNO[t]
                if (n,t,dow) in X and (v(X[n,t,dow]) or 0) > 0.5)
        horas_tipo[n] = round(h, 1)

    # Cobertura real
    cobertura = {}
    for s in slots_activos:
        pab    = req_slot[s]
        req_eu = max(1, -(-pab//2))
        eu_d, aseo_d = [], []
        for dow in DIAS_LAB:
            eu_c   = sum(1 for n in eu_nombres   for t in T_OPTS
                         if (n,t,dow) in X and (v(X[n,t,dow]) or 0) > 0.5
                         and cubre(t,s))
            aseo_c = sum(1 for n in aseo_nombres for t in T_OPTS
                         if (n,t,dow) in X and (v(X[n,t,dow]) or 0) > 0.5
                         and cubre(t,s))
            eu_d.append(eu_c); aseo_d.append(aseo_c)
        cobertura[slot_str(s)] = {
            'pabellones':   pab,
            'req_eu':       req_eu,
            'cob_eu':       min(eu_d),
            'deficit_eu':   max(0, req_eu - min(eu_d)),
            'req_aseo':     1,
            'cob_aseo':     min(aseo_d),
            'deficit_aseo': max(0, 1 - min(aseo_d)),
        }

    # Plantilla mensual
    plantilla = {p['nombre']: {f: semana_tipo[p['nombre']].get(dow,'')
                               for _,sem in enumerate(semanas) for f,dow in sem}
                 for p in personal}
    todas_fechas = sorted({f for _,sem in enumerate(semanas) for f,_ in sem})

    # Turnos para el frontend
    turnos_opt = {t: {
        'ini_slot': TURNOS[t]['ini'],
        'fin_slot': TURNOS[t]['fin'],
        'ini_str':  slot_str(TURNOS[t]['ini']),
        'fin_str':  slot_str(TURNOS[t]['fin']),
        'dur_h':    h_turno(t),
    } for t in T_OPTS}

    return {
        'status':        'ok',
        'solver_status': status,
        'semana_tipo':   semana_tipo,
        'plantilla':     plantilla,
        'horas_tipo':    horas_tipo,
        'cobertura':     cobertura,
        'turnos_opt':    turnos_opt,
        'hora_apertura': slot_str(AP),
        'hora_cierre':   slot_str(CI),
        'semanas':       [{'semana': w+1,
                           'fechas': [f for f,d in sem if d != 'domingo'],
                           'dias':   [(f,d) for f,d in sem]}
                          for w, sem in enumerate(semanas)],
        'fechas':        todas_fechas,
        'personal':      personal,
    }


@hosp_bp.route('/hosp/solve', methods=['POST'])
def hosp_solve():
    try:
        return jsonify(resolver_hosp(request.get_json()))
    except Exception as e:
        return jsonify({'status':'error','mensaje':str(e),
                        'trace':traceback.format_exc()}), 500

@hosp_bp.route('/hosp/ping')
def hosp_ping():
    return jsonify({'status':'ok','modulo':'hospitalizado','version':'4.0'})

@hosp_bp.route('/hosp/template')
def hosp_template():
    """Genera y descarga un template Excel para cargar datos de pabellon."""
    import io
    try:
        import openpyxl
        from flask import send_file
        wb = openpyxl.Workbook()

        # ── Hoja Personal ──────────────────────────────────────────────────────
        ws_p = wb.active; ws_p.title = 'Personal'
        ws_p.append(['Nombre', 'Rol', 'Horas_semana'])
        roles_validos = 'aux_aseo | enfermera_eu | arsenalera | pabellonera | aux_anestesia'
        for row in [
            ('M. DAZA',      'aux_aseo',      42),
            ('F. MONTOYA',   'aux_aseo',      42),
            ('D. CHAVARRIA', 'enfermera_eu',  42),
            ('D. MONSALVE',  'enfermera_eu',  42),
        ]:
            ws_p.append(row)
        ws_p.append([])
        ws_p.append([f'Roles válidos: {roles_validos}'])
        ws_p.column_dimensions['A'].width = 22
        ws_p.column_dimensions['B'].width = 18
        ws_p.column_dimensions['C'].width = 14

        # ── Hoja Demanda ───────────────────────────────────────────────────────
        ws_d = wb.create_sheet('Demanda')
        ws_d.append(['Inicio', 'Fin', 'Pabellones'])
        ws_d.append(['07:00', '10:00', 1])
        ws_d.append(['10:00', '15:00', 2])
        ws_d.append(['15:00', '18:00', 3])
        ws_d.append(['18:00', '21:00', 4])
        ws_d.append(['21:00', '24:00', 2])
        ws_d.column_dimensions['A'].width = 10
        ws_d.column_dimensions['B'].width = 10
        ws_d.column_dimensions['C'].width = 12

        # ── Hoja Configuracion ─────────────────────────────────────────────────
        ws_c = wb.create_sheet('Configuracion')
        ws_c.append(['Parametro', 'Valor'])
        for row in [
            ('mes',              3),
            ('anio',          2025),
            ('horas_max_semana', 42),
            ('hora_apertura', '07:00'),
            ('hora_cierre',   '24:00'),
        ]:
            ws_c.append(row)
        ws_c.column_dimensions['A'].width = 20
        ws_c.column_dimensions['B'].width = 12

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='Template_Pabellon.xlsx')
    except ImportError:
        return jsonify({'status': 'error', 'mensaje': 'openpyxl no instalado. Ejecuta: pip install openpyxl'}), 500