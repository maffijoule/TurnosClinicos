"""
CEM Capacity Planner — Imagen Blueprint
Rutas: /imagen/template
"""
from flask import Blueprint, send_file, jsonify, request
import io
import traceback

imagen_bp = Blueprint('imagen', __name__)

@imagen_bp.route('/imagen/ping')
def imagen_ping():
    return jsonify({'status': 'ok', 'module': 'imagen'})

@imagen_bp.route('/imagen/solve', methods=['POST'])
def imagen_solve():
    try:
        data       = request.get_json()
        tecnologos = data.get('tecnologos', [])
        result     = resolver_semanal_imagen(
            tecnologos         = tecnologos,
            demanda            = data.get('demanda', {}),
            configuracion      = data.get('configuracion', {}),
        )
        result['tecnologos_sel'] = tecnologos
        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'mensaje': str(e),
                        'trace': traceback.format_exc()}), 500


def resolver_semanal_imagen(tecnologos, demanda, configuracion):
    """
    Solver semanal para Imagen — no requiere mes/año.
    Optimiza la semana tipo (Lun-Sáb) minimizando déficit de cobertura de TMs.
    demanda: {dow: {'HH:MM': tm_requeridos}}
    """
    import pulp, sys

    cfg = configuracion

    P            = int(cfg.get('puestos_fisicos_max', int(cfg.get('puestos_max', 4))))
    turno_min_sl = int(float(cfg.get('turno_min_horas', cfg.get('turno_min', 4))) * 2)
    turno_max_sl = int(float(cfg.get('turno_max_horas', cfg.get('turno_max', 9))) * 2)
    horas_min    = float(cfg.get('horas_min_ejecutivo', cfg.get('horas_min_ejec', 0)))

    hora_ap      = float(cfg.get('hora_apertura', 7))
    hora_ci      = float(cfg.get('hora_cierre', 21))
    INICIO_OP    = int(hora_ap * 2)
    FIN_OP       = int(hora_ci * 2)

    hora_ap_sat  = float(cfg.get('hora_apertura_sat', hora_ap))
    hora_ci_sat  = float(cfg.get('hora_cierre_sat', hora_ci))
    INICIO_OP_SAT = int(hora_ap_sat * 2)
    FIN_OP_SAT    = int(hora_ci_sat * 2)

    DIAS_LABOR   = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes']
    nombres      = [t['nombre'] for t in tecnologos]
    h_max        = {t['nombre']: float(t.get('horas_semana', 44)) for t in tecnologos}
    N            = len(nombres)
    K            = list(range(P))
    BIG_M        = (max(FIN_OP, FIN_OP_SAT) - min(INICIO_OP, INICIO_OP_SAT)) + turno_max_sl + 2

    def str_slot(s):
        parts = str(s).strip().split(':')
        h, m = int(parts[0]), int(parts[1])
        return h * 2 + m // 30

    def slot_str(s):
        return f"{s//2:02d}:{(s%2)*30:02d}"

    # ── Demand profiles ──────────────────────────────────────────────────────
    dem_wd = {}
    for dow, slots in demanda.items():
        if dow not in DIAS_LABOR: continue
        for hora, val in slots.items():
            s = str_slot(hora)
            if s == FIN_OP: s = FIN_OP - 1
            dem_wd[s] = max(dem_wd.get(s, 0.0), float(val))

    dem_sat = {}
    if 'sábado' in demanda:
        for hora, val in demanda['sábado'].items():
            s = str_slot(hora)
            if s == FIN_OP_SAT: s = FIN_OP_SAT - 1
            v2 = float(val)
            if v2 > 0: dem_sat[s] = v2

    slots_wd  = {s for s, v in dem_wd.items()  if v > 0 and INICIO_OP <= s < FIN_OP}
    slots_sat = {s for s, v in dem_sat.items() if v > 0 and INICIO_OP_SAT <= s < FIN_OP_SAT}

    print(f"[IMG-SOLVER] N={N} P={P} turno {turno_min_sl/2}-{turno_max_sl/2}h "
          f"wd={hora_ap}-{hora_ci} sat={hora_ap_sat}-{hora_ci_sat} "
          f"slots_wd={len(slots_wd)} slots_sat={len(slots_sat)}", flush=True, file=sys.stderr)

    # ── Contadores de nombre → id corto ──────────────────────────────────────
    _cnt = [0]; _cache = {}
    def si(s):
        s = str(s)
        if s not in _cache: _cache[s] = str(_cnt[0]); _cnt[0] += 1
        return _cache[s]

    # ── Problema PuLP ────────────────────────────────────────────────────────
    prob = pulp.LpProblem("IMG_Semanal", pulp.LpMinimize)

    T_entry_wd  = {k: pulp.LpVariable(f"TEW_{k}", lowBound=INICIO_OP,     upBound=FIN_OP-turno_min_sl,     cat='Integer') for k in K}
    T_dur_wd    = {k: pulp.LpVariable(f"TDW_{k}", lowBound=turno_min_sl,  upBound=turno_max_sl,            cat='Integer') for k in K}
    T_used_wd   = {k: pulp.LpVariable(f"TUW_{k}", cat='Binary') for k in K}

    T_entry_sat = {k: pulp.LpVariable(f"TES_{k}", lowBound=INICIO_OP_SAT, upBound=FIN_OP_SAT-turno_min_sl, cat='Integer') for k in K}
    T_dur_sat   = {k: pulp.LpVariable(f"TDS_{k}", lowBound=turno_min_sl,  upBound=turno_max_sl,            cat='Integer') for k in K}
    T_used_sat  = {k: pulp.LpVariable(f"TUS_{k}", cat='Binary') for k in K}

    COL_MIN_SL  = 13   # 6.5 h en slots de 30 min
    Has_col_wd  = {k: pulp.LpVariable(f"HCW_{k}", cat='Binary') for k in K}
    Has_col_sat = {k: pulp.LpVariable(f"HCS_{k}", cat='Binary') for k in K}

    A_wd  = {(n,k): pulp.LpVariable(f"AWD_{si(n)}_{k}", cat='Binary') for n in nombres for k in K}
    W_sat = {n:     pulp.LpVariable(f"WS_{si(n)}",      cat='Binary') for n in nombres}
    A_sat = {(n,k): pulp.LpVariable(f"AS_{si(n)}_{k}",  cat='Binary') for n in nombres for k in K}

    COV_wd  = {(n,k,s): pulp.LpVariable(f"CWD_{si(n)}_{k}_{s}", cat='Binary') for n in nombres for k in K for s in slots_wd}
    COV_sat = {(n,k,s): pulp.LpVariable(f"CS_{si(n)}_{k}_{s}",  cat='Binary') for n in nombres for k in K for s in slots_sat}

    PENALTY = 4
    SLK1_wd  = {s: pulp.LpVariable(f"SWD1_{s}", lowBound=0, upBound=1) for s in slots_wd  if dem_wd.get(s,0)  > 0}
    SLK2_wd  = {s: pulp.LpVariable(f"SWD2_{s}", lowBound=0)            for s in slots_wd  if dem_wd.get(s,0)  > 0}
    SLK1_sat = {s: pulp.LpVariable(f"SST1_{s}", lowBound=0, upBound=1) for s in slots_sat if dem_sat.get(s,0) > 0}
    SLK2_sat = {s: pulp.LpVariable(f"SST2_{s}", lowBound=0)            for s in slots_sat if dem_sat.get(s,0) > 0}

    H_wd  = {n: pulp.LpVariable(f"HWD_{si(n)}",  lowBound=0, upBound=turno_max_sl) for n in nombres}
    H_sat = {n: pulp.LpVariable(f"HSAT_{si(n)}", lowBound=0, upBound=turno_max_sl) for n in nombres}

    HCN_wd  = {(n,k): pulp.LpVariable(f"HCNWD_{si(n)}_{k}",  cat='Binary') for n in nombres for k in K}
    HCN_sat = {(n,k): pulp.LpVariable(f"HCNSAT_{si(n)}_{k}", cat='Binary') for n in nombres for k in K}

    # Objetivo
    prob += (pulp.lpSum(SLK1_wd.values())  + PENALTY*pulp.lpSum(SLK2_wd.values()) +
             pulp.lpSum(SLK1_sat.values()) + PENALTY*pulp.lpSum(SLK2_sat.values())), "MinDeficit"

    # Asignación
    for n in nombres:
        prob += pulp.lpSum(A_wd[n,k]  for k in K) == 1,        f"OneWD_{si(n)}"
        prob += pulp.lpSum(A_sat[n,k] for k in K) == W_sat[n], f"OneSat_{si(n)}"

    # Turno semana
    for k in K:
        for n in nombres: prob += A_wd[n,k] <= T_used_wd[k], f"TUW_{si(n)}_{k}"
        prob += T_entry_wd[k]+T_dur_wd[k] <= FIN_OP    + BIG_M*(1-T_used_wd[k]), f"TEndW_{k}"
        prob += T_entry_wd[k]             >= INICIO_OP - BIG_M*(1-T_used_wd[k]), f"TStartW_{k}"
    for k in K[:-1]: prob += T_entry_wd[k] <= T_entry_wd[k+1]+BIG_M*(1-T_used_wd[k]), f"OrdW_{k}"

    # Turno sábado
    for k in K:
        for n in nombres: prob += A_sat[n,k] <= T_used_sat[k], f"TUS_{si(n)}_{k}"
        prob += T_entry_sat[k]+T_dur_sat[k] <= FIN_OP_SAT    + BIG_M*(1-T_used_sat[k]), f"TEndS_{k}"
        prob += T_entry_sat[k]              >= INICIO_OP_SAT - BIG_M*(1-T_used_sat[k]), f"TStartS_{k}"
    for k in K[:-1]: prob += T_entry_sat[k] <= T_entry_sat[k+1]+BIG_M*(1-T_used_sat[k]), f"OrdS_{k}"

    # Colación semana y sábado
    for k in K:
        prob += T_dur_wd[k]  >= COL_MIN_SL   - BIG_M*(1-Has_col_wd[k]),  f"HCW_lb_{k}"
        prob += T_dur_wd[k]  <= COL_MIN_SL-1 + BIG_M*Has_col_wd[k],      f"HCW_ub_{k}"
        prob += Has_col_wd[k]  <= T_used_wd[k],                            f"HCW_used_{k}"
        prob += T_dur_sat[k] >= COL_MIN_SL   - BIG_M*(1-Has_col_sat[k]), f"HCS_lb_{k}"
        prob += T_dur_sat[k] <= COL_MIN_SL-1 + BIG_M*Has_col_sat[k],     f"HCS_ub_{k}"
        prob += Has_col_sat[k] <= T_used_sat[k],                           f"HCS_used_{k}"

    # Linealizar HCN
    for n in nombres:
        for k in K:
            prob += HCN_wd[n,k]  <= A_wd[n,k];  prob += HCN_wd[n,k]  <= Has_col_wd[k];  prob += HCN_wd[n,k]  >= A_wd[n,k]+Has_col_wd[k]-1
            prob += HCN_sat[n,k] <= A_sat[n,k]; prob += HCN_sat[n,k] <= Has_col_sat[k]; prob += HCN_sat[n,k] >= A_sat[n,k]+Has_col_sat[k]-1

    # Linealizar H_wd / H_sat
    for n in nombres:
        for k in K:
            prob += H_wd[n]  >= T_dur_wd[k]  - BIG_M*(1-A_wd[n,k])
            prob += H_wd[n]  <= T_dur_wd[k]  + BIG_M*(1-A_wd[n,k])
            prob += H_sat[n] >= T_dur_sat[k] - BIG_M*(1-A_sat[n,k])
            prob += H_sat[n] <= T_dur_sat[k] + BIG_M*(1-A_sat[n,k])
        prob += H_sat[n] <= turno_max_sl * W_sat[n]

    # Horas semanales
    for n in nombres:
        col_wd_n  = pulp.lpSum(HCN_wd[n,k]  for k in K)
        col_sat_n = pulp.lpSum(HCN_sat[n,k] for k in K)
        horas_trab = 5*(H_wd[n]-col_wd_n) + (H_sat[n]-col_sat_n)
        prob += horas_trab <= h_max[n]*2
        if horas_min > 0: prob += horas_trab >= horas_min*2

    # Cobertura semana
    for n in nombres:
        for k in K:
            for s in slots_wd:
                c = COV_wd[n,k,s]
                prob += c <= A_wd[n,k]
                prob += T_entry_wd[k] <= s + BIG_M*(1-c)
                prob += T_entry_wd[k]+T_dur_wd[k] >= s+1 - BIG_M*(1-c)
    for s in slots_wd:
        if dem_wd.get(s,0) <= 0: continue
        prob += pulp.lpSum(COV_wd[n,k,s] for n in nombres for k in K) + SLK1_wd[s]+SLK2_wd[s] >= dem_wd[s]

    # Cobertura sábado
    for n in nombres:
        for k in K:
            for s in slots_sat:
                c = COV_sat[n,k,s]
                prob += c <= A_sat[n,k]
                prob += T_entry_sat[k] <= s + BIG_M*(1-c)
                prob += T_entry_sat[k]+T_dur_sat[k] >= s+1 - BIG_M*(1-c)
    for s in slots_sat:
        if dem_sat.get(s,0) <= 0: continue
        prob += pulp.lpSum(COV_sat[n,k,s] for n in nombres for k in K) + SLK1_sat[s]+SLK2_sat[s] >= dem_sat[s]

    # Apertura / Cierre
    apertura_min     = int(cfg.get('apertura_min', 0))
    cierre_min       = int(cfg.get('cierre_min', 0))
    apertura_min_sat = int(cfg.get('apertura_min_sat', 0))
    cierre_min_sat   = int(cfg.get('cierre_min_sat', 0))
    if apertura_min > 0 and INICIO_OP in slots_wd:
        prob += pulp.lpSum(COV_wd[n,k,INICIO_OP] for n in nombres for k in K) >= apertura_min
    if cierre_min > 0 and (FIN_OP-1) in slots_wd:
        prob += pulp.lpSum(COV_wd[n,k,FIN_OP-1] for n in nombres for k in K) >= cierre_min
    if apertura_min_sat > 0 and INICIO_OP_SAT in slots_sat:
        prob += pulp.lpSum(COV_sat[n,k,INICIO_OP_SAT] for n in nombres for k in K) >= apertura_min_sat
    if cierre_min_sat > 0 and (FIN_OP_SAT-1) in slots_sat:
        prob += pulp.lpSum(COV_sat[n,k,FIN_OP_SAT-1] for n in nombres for k in K) >= cierre_min_sat

    # Solve
    print(f"[IMG-SOLVER] {len(prob.variables())} vars · {len(prob.constraints)} constraints", flush=True, file=sys.stderr)
    prob.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=120))
    v = pulp.value
    status = pulp.LpStatus[prob.status]
    print(f"[IMG-SOLVER] Status: {status}", flush=True, file=sys.stderr)

    has_inc = any(v(T_used_wd[k]) is not None for k in K)
    if prob.status == pulp.constants.LpStatusInfeasible or not has_inc:
        return {'status': 'infeasible', 'mensaje': f'Sin solución factible ({status}). '
                'Revisa turno_min/max, puestos_max, horas de los TMs.'}

    # ── Extraer resultados ───────────────────────────────────────────────────
    col_offset_wd  = {k: int(round(v(T_dur_wd[k])))//2  for k in K if v(T_used_wd[k])>0.5 and v(Has_col_wd[k])>0.5}
    col_offset_sat = {k: int(round(v(T_dur_sat[k])))//2 for k in K if v(T_used_sat[k])>0.5 and v(Has_col_sat[k])>0.5}

    asig_wd  = {n: next((k for k in K if v(A_wd[n,k])  > 0.5), None) for n in nombres}
    asig_sat = {n: next((k for k in K if v(A_sat[n,k]) > 0.5), None) for n in nombres}

    # Semana genérica (un solo período, sin fechas reales)
    DIAS_SEMANA = [('Lun','lunes'),('Mar','martes'),('Mié','miércoles'),
                   ('Jue','jueves'),('Vie','viernes'),('Sáb','sábado')]
    semana = {'semana': 1, 'dias': DIAS_SEMANA}

    turnos_out = []
    for n in nombres:
        k_wd  = asig_wd.get(n)
        k_sat = asig_sat.get(n)
        ent_wd  = int(round(v(T_entry_wd[k_wd])))   if k_wd  is not None else None
        dur_wd  = int(round(v(T_dur_wd[k_wd])))     if k_wd  is not None else None
        ent_sat = int(round(v(T_entry_sat[k_sat]))) if k_sat is not None else None
        dur_sat = int(round(v(T_dur_sat[k_sat])))   if k_sat is not None else None

        for label, dow in DIAS_SEMANA:
            if dow == 'sábado':
                if v(W_sat[n]) > 0.5 and ent_sat is not None:
                    col_off = col_offset_sat.get(k_sat)
                    worked  = (dur_sat - (1 if col_off is not None else 0)) * 0.5
                    turnos_out.append({'ejecutivo':n,'semana':1,'fecha':label,'dia':dow,'libre':False,
                        'entrada':slot_str(ent_sat),'salida':slot_str(ent_sat+dur_sat),
                        'duracion_h':worked,'colacion':col_off is not None,'colacion_offset':col_off})
                else:
                    turnos_out.append({'ejecutivo':n,'semana':1,'fecha':label,'dia':dow,'libre':True})
            else:
                if ent_wd is not None:
                    col_off = col_offset_wd.get(k_wd)
                    worked  = (dur_wd - (1 if col_off is not None else 0)) * 0.5
                    turnos_out.append({'ejecutivo':n,'semana':1,'fecha':label,'dia':dow,'libre':False,
                        'entrada':slot_str(ent_wd),'salida':slot_str(ent_wd+dur_wd),
                        'duracion_h':worked,'colacion':col_off is not None,'colacion_offset':col_off})
                else:
                    turnos_out.append({'ejecutivo':n,'semana':1,'fecha':label,'dia':dow,'libre':True})

    deficit = (sum(v(sv) or 0 for sv in SLK1_wd.values())  + sum(v(sv) or 0 for sv in SLK2_wd.values()) +
               sum(v(sv) or 0 for sv in SLK1_sat.values()) + sum(v(sv) or 0 for sv in SLK2_sat.values()))

    return {
        'status':            'ok',
        'inicio_op':         INICIO_OP,
        'fin_op':            FIN_OP,
        'solver_status':     status,
        'deficit_cobertura': round(deficit, 2),
        'turnos':            turnos_out,
        'dem_wd_profile':    {slot_str(s): dem_wd[s]          for s in sorted(dem_wd)  if dem_wd[s]  > 0},
        'dem_sat_profile':   {slot_str(s): dem_sat.get(s, 0)  for s in sorted(slots_sat)},
        'semanas':           [semana],
        'semana_tipo':       [],
        'horas_ejecutivo':   {},
        'shift_types':       [],
    }

@imagen_bp.route('/imagen/template')
def imagen_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        DIAS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
        EXAM_TYPES = ['Resonancia', 'Scanner', 'Ecografias', 'Rayos', 'Mamografias', 'Densitometria']
        HORAS = list(range(7, 22))  # 07 a 21

        hdr_fill  = PatternFill('solid', fgColor='1E293B')
        hdr_font  = Font(color='FFFFFF', bold=True, size=10)
        thin      = Side(style='thin', color='CBD5E1')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        center    = Alignment(horizontal='center', vertical='center')

        def style_header(cell, fill=None, font=None):
            cell.fill = fill or hdr_fill
            cell.font = font or hdr_font
            cell.alignment = center
            cell.border = border

        def style_cell(cell):
            cell.alignment = center
            cell.border = border

        # ── Hoja Tecnologos ────────────────────────────────────────────────────
        ws_tec = wb.active
        ws_tec.title = 'Tecnologos'
        ws_tec.column_dimensions['A'].width = 20
        ws_tec.column_dimensions['B'].width = 14

        for col, label in enumerate(['Nombre', 'Horas_mes'], 1):
            c = ws_tec.cell(row=1, column=col, value=label)
            style_header(c)

        tecnologos = [
            ('TM 1', 160),
            ('TM 2', 160),
            ('TM 3', 160),
            ('TM 4', 160),
            ('TM 5', 160),
            ('TM 6', 160),
        ]
        for r, (nombre, horas) in enumerate(tecnologos, 2):
            ws_tec.cell(row=r, column=1, value=nombre).border = border
            ws_tec.cell(row=r, column=2, value=horas).border = border

        # ── Hojas de demanda por tipo de examen ───────────────────────────────
        for exam in EXAM_TYPES:
            ws = wb.create_sheet(title=f'Demanda_{exam}')
            ws.column_dimensions['A'].width = 10
            for i, d in enumerate(DIAS, 1):
                ws.column_dimensions[get_column_letter(i + 1)].width = 12

            # Header row
            c = ws.cell(row=1, column=1, value='Hora')
            style_header(c)
            for ci, dia in enumerate(DIAS, 2):
                c = ws.cell(row=1, column=ci, value=dia)
                style_header(c)

            # Data rows (integer hours)
            for ri, hora in enumerate(HORAS, 2):
                c = ws.cell(row=ri, column=1, value=hora)
                style_cell(c)
                c.font = Font(bold=True, size=9)
                for ci in range(2, len(DIAS) + 2):
                    cell = ws.cell(row=ri, column=ci, value=0)
                    style_cell(cell)

        # ── Hoja Configuracion ────────────────────────────────────────────────
        ws_cfg = wb.create_sheet(title='Configuracion')
        ws_cfg.column_dimensions['A'].width = 28
        ws_cfg.column_dimensions['B'].width = 14

        for col, label in enumerate(['Parametro', 'Valor'], 1):
            c = ws_cfg.cell(row=1, column=col, value=label)
            style_header(c)

        cfg_rows = [
            ('mes', 1),
            ('anio', 2025),
            ('puestos_max', 4),
            ('turno_min', 4),
            ('turno_max', 9),
            ('horas_min_ejec', 20),
            ('hora_apertura', 7),
            ('hora_cierre', 21),
            ('apertura_min', 2),
            ('cierre_min', 2),
            ('hora_apertura_sat', 9),
            ('hora_cierre_sat', 14),
            ('apertura_min_sat', 2),
            ('cierre_min_sat', 2),
        ]
        for ri, (param, val) in enumerate(cfg_rows, 2):
            ws_cfg.cell(row=ri, column=1, value=param).border = border
            c = ws_cfg.cell(row=ri, column=2, value=val)
            c.border = border
            c.alignment = center

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='template_imagen.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500
